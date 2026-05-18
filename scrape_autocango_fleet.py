"""
Autocango.com Fleet section scraper.

The /fleet section publishes wholesale offers (Demo/Rental/Test/Auction
cars, factory direct deals) with downloadable Excel pricelists. Each
offer page has a "Download List" button that triggers a download from
i1.autocango.com/fleet/doc/<hash>.xlsx — a public OSS URL.

This script:
  1. Walks the /fleet listing, collects every ACF offer card
     (brand, title, date, MOQ, description, detail URL).
  2. Opens each /fleet/ACFxxx detail page, captures the xlsx URL by
     listening for the request fired when "Download List" is clicked.
  3. Downloads the xlsx with requests (the URL is public, no auth).
  4. Parses every row with openpyxl into fleet_pricelist.
  5. Upserts an `fleet_offers` row per ACF for the catalog index.

Usage:
    python scrape_autocango_fleet.py

Env: SUPABASE_URL, SUPABASE_KEY
"""

import argparse
import os
import re
import sys
import time
import traceback
from datetime import datetime

import requests

import db as DB

SOURCE = "autocango"
LISTING_URL = "https://www.autocango.com/fleet"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/131.0.0.0 Safari/537.36")


# ── Excel column normalization ───────────────────────────────────────────
# The xlsx has a header row like:
#   Brand Series | 车系 | Model Line | 配置 | Trim | MSRP (CNY) | EXW (USD) | FCA Shanghai
# but column positions sometimes differ. We map labels to canonical fields.
HEADER_MAP = {
    # English headers seen across the 5 xlsx variants
    "brand series":              "brand",
    "series":                    "brand",
    "series english":            "brand",
    "model line":                "model_line",
    "model":                     "model_line",
    "model english":             "model_line",
    "model name":                "model_line",
    "trim":                      "trim",
    "trim/spec":                 "trim",
    "model (trim/spec)":         "trim",
    # Prices — accept variants
    "msrp":                      "msrp_cny",
    "msrp (cny)":                "msrp_cny",
    "msrp (after options)":      "msrp_cny",
    "msrp (10000)":              "msrp_cny_wan",   # in 10000 CNY units
    "msrp\n(元)":                 "msrp_cny",
    "exw":                       "exw_usd",
    "exw (usd)":                 "exw_usd",
    "fca":                       "fca_shanghai_usd",
    "fca shanghai":              "fca_shanghai_usd",
    "price":                     "exw_usd",
    "price usd":                 "exw_usd",
    "pric usd":                  "exw_usd",
    "price\nusd":                "exw_usd",
    # Chinese variants
    "车系":                       "brand_zh",
    "车型":                       "model_line_zh",
    "车型名称":                    "model_line_zh",
    "车辆型号":                    "model_line_zh",
    "车辆配置":                    "trim_zh",
    "车型版本":                    "trim_zh",
    "配置":                       "trim_zh",
    "指导价":                      "msrp_cny",
    "厂商指导价":                  "msrp_cny",
}


# Headers whose label is a bilingual "Chinese (English)" string, e.g.
# "指导价 (MSRP)", "车型 (Model)", "配置 (Trim/Spec)". We strip parentheses
# and the leading Chinese block to retrieve the English label.
def _strip_bilingual(label: str) -> str:
    s = re.sub(r"[（(].*?[)）]", "", label).strip()  # drop "(...)"
    # If still Chinese-only, return as-is
    return s


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _to_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return None
    m = re.search(r"-?\d[\d,]*", str(v))
    if not m:
        return None
    try:
        return int(m.group(0).replace(",", ""))
    except Exception:
        return None


def _excel_header_indices(header_row: list) -> dict:
    """Resolve column index → canonical field name from the xlsx header.

    Handles three layout variants observed on autocango:
      1) "Brand Series | 车系 | Model Line | 配置 | Trim | MSRP | EXW | FCA"
         (Changan-style factory pricelist, English columns paired with
         Chinese ones to the left).
      2) "车型 (Model) | 指导价 (MSRP) | 配置 (Trim/Spec) | ..." — single
         bilingual cell per column. We strip the Chinese prefix and match
         the parenthesised English label.
      3) "Series | Series English | Model | Model English | MSRP (10000)..."
         — explicit zh/en sibling columns labelled in English.
    """
    idx_to_field: dict[int, str] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        raw = str(cell)
        key = _norm(raw)
        if key in HEADER_MAP:
            idx_to_field[i] = HEADER_MAP[key]
            continue
        # "指导价 (MSRP)" → try the English in parens
        en_match = re.search(r"[（(]([A-Za-z][^)）]+)[)）]", raw)
        if en_match:
            en = _norm(en_match.group(1))
            if en in HEADER_MAP:
                idx_to_field[i] = HEADER_MAP[en]
                continue
        # Pure Chinese label
        if re.search(r"[一-鿿]", raw) and not re.search(r"[A-Za-z]", raw):
            if key in HEADER_MAP:
                idx_to_field[i] = HEADER_MAP[key]
                continue
        # Composite headers like "FCA Shanghai/Ningbo/Taicang/Guangzhou(+300)"
        if "fca" in key:
            idx_to_field[i] = "fca_shanghai_usd"
            continue
        if "msrp" in key:
            # Distinguish 10000-CNY units
            if "10000" in key or "万" in raw:
                idx_to_field[i] = "msrp_cny_wan"
            else:
                idx_to_field[i] = "msrp_cny"
            continue
    return idx_to_field


def parse_xlsx(path: str) -> list[dict]:
    """Parse an autocango fleet xlsx into normalized rows.

    Strategy:
      - Find the header row (first row with >= 3 recognized labels).
      - The English label and its Chinese counterpart often sit in adjacent
        cells. We pair them up by inspecting nearby columns.
      - Skip rows that don't have at least a brand and a model_line.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    rows_out: list[dict] = []
    for ws in wb.worksheets:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        # Locate the header row by running the per-cell resolver and
        # counting how many cells map to canonical fields.
        header_idx = None
        for i, row in enumerate(all_rows[:10]):
            mapped = _excel_header_indices(list(row))
            if len(mapped) >= 3:
                header_idx = i
                break
        if header_idx is None:
            continue
        header = list(all_rows[header_idx])
        col_map = _excel_header_indices(header)

        # In two-column-per-field layouts (Chinese cell left of English
        # cell for brand/model_line/trim), attach the *_zh sibling.
        bilingual_pairs = {}
        for idx, field in sorted(col_map.items()):
            if field in ("brand", "model_line", "trim"):
                zh_idx = idx - 1
                if zh_idx >= 0 and zh_idx not in col_map:
                    cell = header[zh_idx]
                    if cell and re.search(r"[一-鿿]", str(cell)):
                        bilingual_pairs[zh_idx] = field + "_zh"
        col_map.update(bilingual_pairs)

        # Some files put zh on the right of en — try the mirror too.
        for idx, field in sorted(col_map.items()):
            if field in ("brand", "model_line", "trim"):
                zh_idx = idx + 1
                if zh_idx < len(header) and zh_idx not in col_map:
                    cell = header[zh_idx]
                    if cell and re.search(r"[一-鿿]", str(cell)) \
                            and not re.search(r"[A-Za-z]", str(cell)):
                        bilingual_pairs[zh_idx] = field + "_zh"
        col_map.update(bilingual_pairs)

        for r_idx in range(header_idx + 1, len(all_rows)):
            row = all_rows[r_idx]
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            rec = {f: None for f in (
                "brand", "brand_zh", "model_line", "model_line_zh",
                "trim", "trim_zh", "msrp_cny", "exw_usd", "fca_shanghai_usd",
            )}
            for idx, field in col_map.items():
                if idx >= len(row):
                    continue
                val = row[idx]
                if field == "msrp_cny_wan":
                    # Stored as "10000 CNY" units in the source
                    n = _to_int(val) if isinstance(val, str) else val
                    if isinstance(n, (int, float)):
                        rec["msrp_cny"] = int(round(float(val) * 10000))
                    continue
                if field.endswith("_cny") or field.endswith("_usd"):
                    rec[field] = _to_int(val)
                else:
                    if val is not None:
                        rec[field] = re.sub(r"\s+", " ", str(val)).strip()
            # Keep rows that have at least one identifying string OR a price.
            has_identity = any(rec.get(k) for k in (
                "brand", "brand_zh", "model_line", "model_line_zh", "trim"))
            has_price = any(rec.get(k) for k in (
                "msrp_cny", "exw_usd", "fca_shanghai_usd"))
            if not (has_identity and has_price):
                continue
            # Serialise raw row so we don't lose VIN-level columns
            # (exterior color, mileage, reg date, options, …).
            rec["raw_row"] = _row_to_json(header, row)
            rows_out.append(rec)
    return rows_out


def _row_to_json(header: list, row: tuple) -> dict:
    """Serialize one xlsx row as {column_label: value} (stringified)."""
    out: dict[str, object] = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        label = header[i] if i < len(header) and header[i] is not None else f"col_{i}"
        label = re.sub(r"\s+", " ", str(label)).strip()
        # JSON-safe value
        if isinstance(cell, datetime):
            val = cell.strftime("%Y-%m-%d")
        elif isinstance(cell, (int, float, str, bool)):
            val = cell
        else:
            val = str(cell)
        out[label] = val
    return out


# ── Listing + detail page scrapers (playwright) ──────────────────────────

def collect_offers(page) -> list[dict]:
    """Scrape the /fleet listing for offer cards.

    Each card is rendered as a <div class="flex-1 min-w-0"> wrapping
    title / date / brand / "ID: ACFxxx" / "MOQ: N" in order, all
    siblings inside one container that also holds the detail-page <a>.
    """
    page.goto(LISTING_URL, wait_until="networkidle", timeout=60_000)
    page.wait_for_timeout(2000)
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(1500)

    offers = page.evaluate(r"""() => {
        // Walk up from each "ID: ACFxxx" leaf to the card container.
        const idNodes = [...document.querySelectorAll('*')]
            .filter(el => el.children.length === 0 &&
                          /^ID:\s*ACF\d+/.test(el.textContent || ''));
        const out = [];
        for (const idn of idNodes) {
            let card = idn;
            for (let k = 0; k < 8; k++) {
                const par = card.parentElement;
                if (!par) break;
                if (par.querySelector("a[href^='/fleet/ACF']") &&
                    par.textContent.length > 80) { card = par; break; }
                card = par;
            }
            const a = card.querySelector("a[href^='/fleet/ACF']");
            const acf = (idn.textContent.match(/ACF\d+/) || [])[0];
            const lines = (card.innerText || '').split('\n')
                .map(s => s.trim()).filter(Boolean);
            // Layout: [title, date(YYYY-MM-DD), brand, "ID: ACFxxx", "MOQ: N", ...desc]
            let title = null, date = null, brand = null, moq = null;
            const descParts = [];
            let pastId = false;
            for (const ln of lines) {
                if (!title && !/^\d{4}-\d{2}-\d{2}$/.test(ln)
                    && !/^ID:/.test(ln) && !/^MOQ:/.test(ln)) { title = ln; continue; }
                if (!date && /^\d{4}-\d{2}-\d{2}$/.test(ln)) { date = ln; continue; }
                if (date && !brand && !/^ID:/.test(ln)) { brand = ln; continue; }
                if (/^MOQ:\s*(\d+)/i.test(ln)) {
                    moq = parseInt(ln.match(/\d+/)[0], 10); pastId = true; continue;
                }
                if (/^ID:/.test(ln)) { pastId = true; continue; }
                if (pastId) descParts.push(ln);
            }
            out.push({
                acf, title, brand, date, moq,
                description: descParts.join(' ').slice(0, 2000),
                url: a ? new URL(a.getAttribute('href'), location.origin).href
                       : `${location.origin}/fleet/${acf}`,
            });
        }
        return out;
    }""")
    return offers


def capture_xlsx_url(page, offer_url: str) -> str | None:
    """Navigate to offer page, click Download List, return the xlsx URL."""
    page.goto(offer_url, wait_until="networkidle", timeout=60_000)
    page.wait_for_timeout(2000)
    # Dismiss cookie banner if shown
    try:
        page.locator("button:has-text('Accept')").click(timeout=2000)
        page.wait_for_timeout(300)
    except Exception:
        pass

    captured: dict[str, str] = {}

    def on_request(r):
        if ".xlsx" in r.url and "/fleet/doc/" in r.url:
            captured["url"] = r.url.split("?")[0]

    page.on("request", on_request)

    try:
        # Listen for the download event AND the xlsx URL request.
        try:
            with page.expect_download(timeout=10_000):
                page.locator("button:has-text('Download List')").first.click()
        except Exception:
            # Some clicks open in a new context — still pick up via on_request
            pass
        page.wait_for_timeout(2000)
    finally:
        page.remove_listener("request", on_request)

    return captured.get("url")


def download_xlsx(url: str, dest: str) -> bool:
    try:
        r = requests.get(url, timeout=60, headers={"User-Agent": UA})
        if r.status_code != 200:
            return False
        with open(dest, "wb") as f:
            f.write(r.content)
        return True
    except Exception as e:
        print(f"  download FAIL: {e}")
        return False


# ── DB upserts ───────────────────────────────────────────────────────────

def upsert_offer(rec: dict) -> bool:
    if not DB.USE_POSTGRES:
        return True  # SQLite has no fleet tables — skip
    payload = {
        "source": SOURCE,
        "source_offer_id": rec["source_offer_id"],
        "brand": rec.get("brand"),
        "title": rec.get("title"),
        "description": rec.get("description"),
        "posted_date": rec.get("posted_date"),
        "moq": rec.get("moq"),
        "url": rec.get("url"),
        "xlsx_url": rec.get("xlsx_url"),
        "rows_count": rec.get("rows_count"),
        "last_seen": DB.now_iso(),
    }
    r = DB._pg_request(
        "POST",
        "fleet_offers?on_conflict=source,source_offer_id",
        headers={"Prefer": "return=minimal,resolution=merge-duplicates"},
        json=payload,
    )
    return r.status_code in (200, 201, 204)


def upsert_pricelist_rows(offer_id: str, xlsx_url: str, rows: list[dict]) -> int:
    if not DB.USE_POSTGRES or not rows:
        return 0
    ok = 0
    # PostgREST handles bulk upserts with arrays — send chunks of 50.
    payload = [
        {
            "source": SOURCE,
            "source_offer_id": offer_id,
            "brand": r.get("brand"),
            "brand_zh": r.get("brand_zh"),
            "model_line": r.get("model_line"),
            "model_line_zh": r.get("model_line_zh"),
            "trim": r.get("trim"),
            "trim_zh": r.get("trim_zh"),
            "msrp_cny": r.get("msrp_cny"),
            "exw_usd": r.get("exw_usd"),
            "fca_shanghai_usd": r.get("fca_shanghai_usd"),
            "xlsx_url": xlsx_url,
            "raw_row": r.get("raw_row"),
            "last_seen": DB.now_iso(),
        }
        for r in rows
    ]
    for start in range(0, len(payload), 50):
        chunk = payload[start:start + 50]
        r = DB._pg_request(
            "POST",
            ("fleet_pricelist?on_conflict=source,source_offer_id,brand,"
             "model_line,trim"),
            headers={"Prefer": "return=minimal,resolution=merge-duplicates"},
            json=chunk,
        )
        if r.status_code in (200, 201, 204):
            ok += len(chunk)
        else:
            print(f"  PG insert FAIL {r.status_code}: {r.text[:300]}")
    return ok


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--keep-xlsx", action="store_true",
                    help="Keep downloaded xlsx files in /tmp")
    args = ap.parse_args()

    from playwright.sync_api import sync_playwright

    print(f"Backend: {DB.backend_name()}")
    print(f"Source: {SOURCE} (section=fleet)\n")

    started = time.time()
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled",
                  "--no-sandbox", "--disable-dev-shm-usage",
                  "--ignore-certificate-errors"],
        )
        ctx = browser.new_context(user_agent=UA,
                                  viewport={"width": 1366, "height": 768},
                                  locale="en-US",
                                  ignore_https_errors=True,
                                  accept_downloads=True)
        ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        page = ctx.new_page()

        offers = collect_offers(page)
        print(f"Collected {len(offers)} offers\n")
        if not offers:
            print("Nothing found — autocango layout may have changed.")
            return

        total_rows = 0
        ok_offers = 0
        for i, offer in enumerate(offers, 1):
            acf = offer["acf"]
            print(f"[{i}/{len(offers)}] {acf} — {offer.get('brand')}: "
                  f"{(offer.get('title') or '')[:60]}")
            xlsx_url = capture_xlsx_url(page, offer["url"])
            if not xlsx_url:
                print(f"    no xlsx URL captured, skipping")
                continue
            dest = f"/tmp/fleet_{acf}.xlsx"
            if not download_xlsx(xlsx_url, dest):
                continue
            try:
                rows = parse_xlsx(dest)
            except Exception as e:
                print(f"    parse FAIL: {e}")
                continue
            print(f"    xlsx -> {len(rows)} priced rows")

            offer_rec = {
                "source_offer_id": acf,
                "brand": offer.get("brand"),
                "title": offer.get("title"),
                "description": offer.get("description"),
                "posted_date": offer.get("date"),
                "moq": offer.get("moq"),
                "url": offer.get("url"),
                "xlsx_url": xlsx_url,
                "rows_count": len(rows),
            }
            if args.dry_run:
                print(f"    DRY-RUN — first row: {rows[0] if rows else None}")
            else:
                upsert_offer(offer_rec)
                wrote = upsert_pricelist_rows(acf, xlsx_url, rows)
                print(f"    upserted {wrote}/{len(rows)} pricelist rows")
                total_rows += wrote
                ok_offers += 1

            if not args.keep_xlsx:
                try:
                    os.remove(dest)
                except Exception:
                    pass

        browser.close()

    elapsed = int(time.time() - started)
    print(f"\nDone in {elapsed}s. Offers: {ok_offers}/{len(offers)}, "
          f"pricelist rows: {total_rows}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        sys.exit(1)
